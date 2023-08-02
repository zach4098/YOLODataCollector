import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import params

map = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'W', 
       'X', 'Y', 'Z']

labelMap = ["pedestrian",     "bicycle",    "car",    "motorbike",    "aeroplane",    "bus",    "train",
            "truck",          "boat"]

def ParseLabelList(text):
    text = text.split(', ')
    whiteList = []
    for i in text:
        if i == "all":
            whiteList.append("all")
        elif i in labelMap:
            whiteList.append(labelMap.index(i))
        elif i not in labelMap:
            print("could not find {} in the label map!".format(i))
    return whiteList

def mainScript(folder, file):
    with open("./{}/{}".format(folder, file), "r") as f:
        lines = f.readlines()
    timeInput = input("Enter Time Recording Started (XX:XX:XX): ").replace(':', '')
    if timeInput == "":
        timeInput = "000000"
    timeInit = ReadRawTime(timeInput)
    startTime, endTime, timeDifference, timeElapsed = TimeDifference(lines, init = timeInit)
    global adjusterStart, adjusterEnd
    adjusterStart = timedelta(days=0, hours=startTime.hour,minutes=startTime.minute,seconds=startTime.second)
    adjusterEnd = timedelta(days=0, hours=timeInit.hour, minutes=timeInit.minute,seconds=timeInit.second)
    timeAdjustedStart = timeInit + adjusterStart
    timeAdjustedEnd = endTime + adjusterEnd

    morningPeakVehicles, nightPeakVehicles, morningPeakStart, morningPeakEnd, nightPeakStart, nightPeakEnd = InitializePeak(lines)

    vehicleChoice = input("Choose Vehicles(s) to Record: ")
    if vehicleChoice == 'all':
        ReadFile(folder, file, lines, timeAdjustedStart, timeAdjustedEnd, timeDifference, morningPeakVehicles, nightPeakVehicles, morningPeakStart, morningPeakEnd, nightPeakStart, nightPeakEnd)
    else:
        reads = ParseLabelList(vehicleChoice)
        for item in reads:
            if item == "all":
                ReadFile(folder, file, lines, timeAdjustedStart, timeAdjustedEnd, timeDifference, morningPeakVehicles, nightPeakVehicles, morningPeakStart, morningPeakEnd, nightPeakStart, nightPeakEnd)
            else:
                ReadFile(folder, file, lines, timeAdjustedStart, timeAdjustedEnd, timeDifference, morningPeakVehicles, nightPeakVehicles, morningPeakStart, morningPeakEnd, nightPeakStart, nightPeakEnd, vType = item)

def ReadFile(folder, file, lines, startTime, endTime, differenceTime, mPeakV, nPeakV, mPeakS, mPeakE, nPeakS, nPeakE, vType=None,):
    count = 0
    for item in lines:
        lines[count] = item.replace("\n", "")
        count += 1
    timeAdjustedStart, timeAdjustedEnd, timeDifference = startTime, endTime, differenceTime
    vehiclesPerHour, totalHours = CalculatePerHours(lines=lines, returnTotalHours=True)
    if vType is not None:
        lines = FilterLines(lines, vType)
        totalVehicles, vFromLeft, vFromRight, vFromNull = TotalVehicles(lines)
    if vType is None:
        totalVehicles, vFromLeft, vFromRight, vFromNull = TotalVehicles(lines)
    vehiclesPerHour = CalculatePerHours(lines=lines, total_hours=totalHours, returnTotalHours=False)
    vPerHourLeft = CalculatePerHours(lines=lines, total_hours=totalHours, numVehicles=vFromLeft)
    vPerHourRight = CalculatePerHours(lines=lines, total_hours=totalHours, numVehicles=vFromRight)

    morningPeakVehicles, nightPeakVehicles, morningPeakStart, morningPeakEnd, nightPeakStart, nightPeakEnd = mPeakV, nPeakV, mPeakS, mPeakE, nPeakS, nPeakE
    if vType is not None:
        if morningPeakVehicles != "N/A":
            morningPeakVehicles = FilterLines(morningPeakVehicles, vType)
        if nightPeakVehicles != "N/A":
            nightPeakVehicles = FilterLines(nightPeakVehicles, vType)
    if morningPeakVehicles != "N/A":
        mPeakTotal, mPeakFromLeft, mPeakFromRight, mPeakFromNull = TotalVehicles(morningPeakVehicles)
        if len(morningPeakVehicles) > 1:
            MvPerHour, mPeakTotalHours = CalculatePerHours(lines=morningPeakVehicles, returnTotalHours=True)
            mvPerHourLeft = CalculatePerHours(lines=morningPeakVehicles, numVehicles=mPeakFromLeft)
            mVPerHourRight = CalculatePerHours(lines=morningPeakVehicles, numVehicles=mPeakFromRight)
        else:
            mPeakTotal = mPeakFromLeft = mPeakFromRight = mPeakFromNull = MvPerHour = mvPerHourLeft = mVPerHourRight = mPeakTotalHours = "N/A"
    else:
        mPeakTotal = mPeakFromLeft = mPeakFromRight = mPeakFromNull = MvPerHour = mvPerHourLeft = mVPerHourRight = mPeakTotalHours = "N/A"

    if nightPeakVehicles != "N/A":
        nPeakTotal, nPeakFromLeft, nPeakFromRight, nPeakFromNull = TotalVehicles(nightPeakVehicles)
        if len(nightPeakVehicles)  >  1:
            NvPerHour, nPeakTotalHours = CalculatePerHours(lines=nightPeakVehicles, returnTotalHours=True)
            nvPerHourLeft = CalculatePerHours(lines=nightPeakVehicles, numVehicles=nPeakFromLeft)
            nVPerHourRight = CalculatePerHours(lines=nightPeakVehicles, numVehicles=nPeakFromRight)
        else:
            nPeakTotal = nPeakFromLeft = nPeakFromRight = nPeakFromNull = NvPerHour = nvPerHourLeft = nVPerHourRight = nPeakTotalHours = "N/A"
    else:
        nPeakTotal = nPeakFromLeft = nPeakFromRight = nPeakFromNull = NvPerHour = nvPerHourLeft = nVPerHourRight = nPeakTotalHours = "N/A"

    if morningPeakVehicles != "N/A":
        morningPeakStart, morningPeakEnd = RefreshPeakTimes(morningPeakVehicles)
    if nightPeakVehicles != "N/A":
        nightPeakStart, nightPeakEnd = RefreshPeakTimes(nightPeakVehicles)

    wb = load_workbook("DataSpread/{}".format(params.defaultSpreadsheetName))
    ws = wb.active
    currentColumn = 2
    openRow = False

    while not openRow:
        if str(ws.cell(1, currentColumn).value) == "None":
            openRow = True
        else:
            currentColumn += 1
    ws.column_dimensions[map[currentColumn - 1]].width = 15
    ws.cell(1, currentColumn).value = file
    ws.cell(2, currentColumn).value = labelMap[vType] if vType is not None else "all"
    ws.cell(3, currentColumn).value = str(timeAdjustedStart.time())
    ws.cell(4, currentColumn).value = str(timeAdjustedEnd.time())
    ws.cell(5, currentColumn).value = str(timeDifference)
    ws.cell(6, currentColumn).value = totalVehicles
    ws.cell(7, currentColumn).value = vFromLeft
    ws.cell(8, currentColumn).value = vFromRight
    ws.cell(9, currentColumn).value = totalHours
    ws.cell(10, currentColumn).value = vehiclesPerHour
    ws.cell(11, currentColumn).value = vPerHourLeft
    ws.cell(12, currentColumn).value = vPerHourRight
    ws.cell(13, currentColumn).value = str(morningPeakStart)
    ws.cell(14, currentColumn).value = str(morningPeakEnd)
    ws.cell(15, currentColumn).value = mPeakTotal
    ws.cell(16, currentColumn).value = mPeakTotalHours
    ws.cell(17, currentColumn).value = MvPerHour
    ws.cell(18, currentColumn).value = mPeakFromLeft
    ws.cell(19, currentColumn).value = mPeakFromRight
    ws.cell(20, currentColumn).value = mvPerHourLeft
    ws.cell(21, currentColumn).value = mVPerHourRight
    ws.cell(22, currentColumn).value = str(nightPeakStart)
    ws.cell(23, currentColumn).value = str(nightPeakEnd)
    ws.cell(24, currentColumn).value = nPeakTotal
    ws.cell(25, currentColumn).value = nPeakTotalHours
    ws.cell(26, currentColumn).value = NvPerHour
    ws.cell(27, currentColumn).value = nPeakFromLeft
    ws.cell(28, currentColumn).value = nPeakFromRight
    ws.cell(29, currentColumn).value = nvPerHourLeft
    ws.cell(30, currentColumn).value = nVPerHourRight

    wb.save("DataSpread/{}".format(params.defaultSpreadsheetName))

def FilterLines(lines, vehicle_type):
    count = 0
    removalList = []
    for item in lines:
        vehicleType = ReadItemType(item)
        if vehicleType != vehicle_type:
            removalList.append(lines[count])
        count += 1

    lines = [i for i in lines if i not in removalList]
    
    return lines


def ReadItemType(textItem):
    Item = textItem.split(',')[3]
    return int(Item)

def ReadItemTime(textItem):
    rawTime = textItem.split(',')[1]
    Time = datetime.strptime(f"{rawTime[:2]}:{rawTime[2:4]}:{rawTime[4:]}", "%H:%M:%S")
    return Time

def ReadRawTime(rawTime):
    if len(rawTime) < 6:
        x = "0"
        rawTime = x + rawTime
    Time = datetime.strptime(f"{rawTime[:2]}:{rawTime[2:4]}:{rawTime[4:]}", "%H:%M:%S")
    return Time

def TimeDifference(lines, **kwargs):
    timeInit = ReadItemTime(lines[0])
    timeFinal = ReadItemTime(lines[len(lines) - 1])
    timeDifference = timeFinal - timeInit
    timeDifferenceRaw = timedelta(days=0, hours=timeFinal.hour-timeInit.hour,minutes=timeFinal.minute-timeInit.minute,
                                  seconds=timeFinal.second-timeInit.second)
    for key, value in kwargs.items():
        if key == "init":
            elapsedTime = value + timeDifferenceRaw
            return timeInit, timeFinal, timeDifferenceRaw, elapsedTime.time()
    return timeInit, timeFinal, timeDifferenceRaw.time()

def TotalVehicles(lines, *types):
    count = 0
    leftCount = 0
    rightCount = 0
    nullCount = 0
    if len(types) != 0:
        for item in lines:
            direction = int(item.split(',')[2])
            vType = int(item.split(',')[3])
            if vType == types[0]:
                count += 1
                if direction == 0:
                    leftCount += 1
                elif direction == 1:
                    rightCount += 1
                else:
                    nullCount += 1
        return count, leftCount, rightCount, nullCount
    else:
        for item in lines:
            count += 1
            direction = int(item.split(',')[2])
            if direction == 0:
                leftCount += 1
            elif direction == 1:
                rightCount += 1
            else:
                nullCount += 1
        return count, leftCount, rightCount, nullCount

def PeakHours(timeInit, timeFinal, lines):
    vehicles = []
    for item in lines:
        vTime = ReadItemTime(item)
        vTime = vTime + adjusterEnd
        if timeInit <= vTime.time() <= timeFinal:
            vehicles.append(item)
    peakStart = ReadItemTime(vehicles[0])
    peakEnd = ReadItemTime(vehicles[len(vehicles) - 1])
    return peakStart, peakEnd, vehicles

def CalculatePerHours(lines, total_hours = None, numVehicles = None, returnTotalHours = False):
    if numVehicles is None:
        vehicleCount = len(lines)
    else:
        vehicleCount = numVehicles
    
    if returnTotalHours or total_hours == None:
        timeInit = ReadItemTime(lines[0])
        timeFinal = ReadItemTime(lines[len(lines) - 1])
        timeDelta = timeFinal - timeInit
        timeBlank = datetime(1900, 1, 1, 0, 0, 0)
        timeDelta = timeBlank + timeDelta
        """timeDelta = datetime(year = 1900, month=1, day=1, hour=timeFinal.hour-timeInit.hour,
                            minute=timeFinal.minute-timeInit.minute,
                            second=timeFinal.second-timeInit.second)"""
        hourDelta = timeDelta.hour+timeDelta.minute/60+timeDelta.second/3600

    if not returnTotalHours and total_hours is not None:
        hourDelta = total_hours
    
    vehiclesPerHour = vehicleCount/hourDelta
    if returnTotalHours:
        return vehiclesPerHour, hourDelta
    if not returnTotalHours:
        return vehiclesPerHour

def InitializePeak(lines):
    returnPeakM = False
    returnPeakN = False
    peakHoursM = input("Morning Peak? y/n: ")
    if peakHoursM == "y":
        returnPeakM = True
        mPeakHoursStart = input("Input Peak Start: ")
        if mPeakHoursStart == "":
            mPeakHoursStart = ReadRawTime(params.defaultMorningPeakStart)
            mPeakHoursStart = mPeakHoursStart.time()
        elif mPeakHoursStart != "":
            mPeakHoursStart = mPeakHoursStart.split(':')
            if len(mPeakHoursStart[0]) == 1:
                mPeakHoursStart[0] = "0" + mPeakHoursStart[0]
            if len(mPeakHoursStart) == 2:
                mPeakHoursStart.append("00")
            mPeakHoursStart = ReadRawTime("".join(mPeakHoursStart))
            mPeakHoursStart = mPeakHoursStart.time()
        mPeakHoursEnd = input("Input Peak End: ")
        if mPeakHoursEnd == "":
            mPeakHoursEnd = ReadRawTime(params.defaultMorningPeakEnd)
            mPeakHoursEnd = mPeakHoursEnd.time()
        elif mPeakHoursEnd != "":
            mPeakHoursEnd = mPeakHoursEnd.split(':')
            if len(mPeakHoursEnd[0]) == 1:
                mPeakHoursEnd[0] = "0" + mPeakHoursEnd[0]
            if len(mPeakHoursEnd) == 2:
                mPeakHoursEnd.append("00")
            mPeakHoursEnd = ReadRawTime("".join(mPeakHoursEnd))
            mPeakHoursEnd = mPeakHoursEnd.time()

        mPeakRealStart, mPeakRealEnd, morningPeakVehicles = PeakHours(mPeakHoursStart, mPeakHoursEnd, lines)
        mPeakRealStart += adjusterEnd
        mPeakRealEnd = mPeakRealEnd + adjusterEnd
    peakHoursN = input("Night Peak? y/n: ")
    if peakHoursN == 'y':
        returnPeakN = True
        nPeakHoursStart = input("Input Peak Start: ")
        if nPeakHoursStart == "":
            nPeakHoursStart = ReadRawTime(params.defaultNightPeakStart)
            nPeakHoursStart = nPeakHoursStart.time()
        elif nPeakHoursStart != "":
            nPeakHoursStart = nPeakHoursStart.split(":")
            if len(nPeakHoursStart[0]) == 1:
                nPeakHoursStart[0] = "0" + nPeakHoursStart[0]
            if len(nPeakHoursStart) == 2:
                nPeakHoursStart.append("00")
            nPeakHoursStart = ReadRawTime("".join(nPeakHoursStart))
            nPeakHoursStart = nPeakHoursStart.time()

        nPeakHoursEnd = input("Input Peak End: ")
        if nPeakHoursEnd == "":
            nPeakHoursEnd = ReadRawTime(params.defaultNightPeakEnd)
            nPeakHoursEnd = nPeakHoursEnd.time()
        elif nPeakHoursEnd != "":
            nPeakHoursEnd = nPeakHoursEnd.split(":")
            if len(nPeakHoursEnd[0]) == 1:
                nPeakHoursEnd[0] = "0" + nPeakHoursEnd[0]
            if len(nPeakHoursEnd) == 2:
                nPeakHoursEnd.append("00")
            nPeakHoursEnd = ReadRawTime("".join(nPeakHoursEnd))
            nPeakHoursEnd = nPeakHoursEnd.time()
        nPeakRealStart, nPeakRealEnd, nightPeakVehicles = PeakHours(nPeakHoursStart, nPeakHoursEnd, lines)
        nPeakRealStart = nPeakRealStart + adjusterEnd
        nPeakRealEnd = nPeakRealEnd + adjusterEnd
    
    if not returnPeakM:
        morningPeakVehicles = mPeakHoursStart = mPeakHoursEnd = mPeakRealStart = mPeakRealEnd = "N/A"
        return morningPeakVehicles, nightPeakVehicles, mPeakRealStart, mPeakRealEnd, nPeakRealStart.time(), nPeakRealEnd.time()
    if not returnPeakN:
        nightPeakVehicles = nPeakHoursStart = nPeakHoursEnd =  nPeakRealStart = nPeakRealEnd ="N/A"
        return morningPeakVehicles, nightPeakVehicles, mPeakRealStart.time(), mPeakRealEnd.time(), nPeakRealStart, nPeakRealEnd

    else:
        return morningPeakVehicles, nightPeakVehicles, mPeakRealStart.time(), mPeakRealEnd.time(), nPeakRealStart.time(), nPeakRealEnd.time()

def RefreshPeakTimes(lines):
    if len(lines) > 1:
        peakStart = ReadItemTime(lines[0])
        peakEnd = ReadItemTime(lines[len(lines) - 1])
        peakStart += adjusterEnd
        peakEnd += adjusterEnd
        return peakStart.time(), peakEnd.time()
    else:
        peakStart = "N/A"
        peakEnd = "N/A"
        return peakStart, peakEnd
    

#MAIN

loop = True
while loop:
    actionInput = input("Select Action: ")
    if actionInput == "clear":
        with open("ClearSpreadsheet.py") as f:
            exec(f.read())
    elif actionInput == "new":
        with open("NewSpreadsheet.py") as f:
            exec(f.read())
    elif actionInput == "read":
        folder = params.FolderName
        collections = os.listdir("{}/".format(folder))

        msg = "Select File to Read:\n"
        count = 0
        for item in collections:
            msg = msg + "{}: {}".format(count, str(item)) + "\n"
            count += 1
        print(msg)
        readLoop = True
        while readLoop:
            fileInput = input("Select File: ")

            if fileInput == "all":
                count = 0
                for i in collections:
                    mainScript(folder, collections[count])
                    count += 1
                print("Added a total of {} datasets!".format(count))
                loop = False
            elif fileInput == "end" or fileInput == "":
                readLoop = False
            else:
                file = collections[int(fileInput)]
                mainScript(folder, file)
                print("Done!")
    elif actionInput == "" or actionInput == "end":
        print("ending!")
        loop = False
    elif actionInput == "help":
        print("Choose one of the following commands:\nread\nclear\nnew\nend")
    else:
        print("Error: no action found!")
        loop = False