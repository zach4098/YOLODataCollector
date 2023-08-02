import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import params

map = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'W', 
       'X', 'Y', 'Z']

def ReadTime(text):
    time = text.split('-')[1].split(' ')[3].split(':')
    return time
def TimeDifference(timeInit, timeFinal, vehicleCount, leftCount, rightCount):
    if vehicleCount == "N/A":
        differenceInHours = differenceInMinutes = differenceInSeconds = totalHours = VperHour = VperHrL = VperHrR = "N/A"
    else:
        initHour = int(timeInit[0])
        initMinute = int(timeInit[1])
        initSecond = int(timeInit[2])

        finalHour = int(timeFinal[0])
        finalMinute = int(timeFinal[1])
        finalSecond = int(timeFinal[2])

        differenceInHours = finalHour-initHour
        differenceInMinutes = finalMinute - initMinute
        differenceInSeconds = finalSecond - initSecond

        if differenceInMinutes < 0:
            differenceInHours -= 1
            differenceInMinutes += 60
        if differenceInSeconds < 0:
            differenceInMinutes -= 1
            differenceInSeconds += 60
        
        totalHours = differenceInHours + (differenceInMinutes/60) + (differenceInSeconds/3600)
        totalHours = round(totalHours, 5)
        
        VperHour = round(vehicleCount/totalHours, 2)
        VperHrL = round(leftCount/totalHours, 2)
        VperHrR = round(rightCount/totalHours, 2)


    return differenceInHours, differenceInMinutes, differenceInSeconds, totalHours, VperHour, VperHrL, VperHrR
def CountLeftRight(List):
    leftCount = 0
    rightCount = 0
    nullCount = 0
    for item in List:
        direction = item.split('-')[2]
        if direction == "left":
            leftCount += 1
        elif direction == "right":
            rightCount += 1
        else:
            nullCount += 1
    return leftCount, rightCount, nullCount

def PeakHours(timeInit, timeFinal, folder, file):
    mPeak = input("Morning Peak? y/n: ")
    if mPeak == 'y':
        mPeakS = input("Enter Morning Peak Start Hours in format XX:XX - ")
        mPeakF = input("Enter Morning Peak End Hours in format XX:XX - ")
        mPeakS = mPeakS.split(':')
        mPeakF = mPeakF.split(':')
        mPeakS = [int(i) for i in mPeakS]
        mPeakF = [int(i) for i in mPeakF]
        MorningPeakHours = [mPeakS[0], mPeakS[1], 0, mPeakF[0], mPeakF[1], 0]
        peakHoursM = True
    elif mPeak == '':
        MorningPeakHours = params.defaultMorningPeakHours
        peakHoursM = True
    else:
        peakHoursM = False
    nPeak = input("Night Peak? y/n: ")
    if nPeak == 'y':
        nPeakS = input("Enter Night Peak Start Hours in format XX:XX - ")
        nPeakF = input("Enter Night Peak End Hours in format XX:XX - ")
        nPeakS = nPeakS.split(":")
        nPeakF = nPeakF.split(":")
        nPeakS = [int(i) for i in nPeakS]
        nPeakF = [int(i) for i in nPeakF]
        NightPeakHours = [nPeakS[0], nPeakS[1], 0, nPeakF[0], nPeakF[1], 0]
        peakHoursN = True
    elif nPeak == '':
        NightPeakHours = params.defaultNightPeakHours
        peakHoursN = True
    else:
        peakHoursN = False
    if peakHoursM:
        count = 0
        hoursFound = False
        with open("{}/{}".format(folder, file), "r") as f:
            lines = f.readlines()
        for item in lines:
            item = item.replace("\n", "")
            item = item.replace("Vehicle ", "")
            lines[count] = item
            count += 1
        count = 0
        start = False
        while not start:
            try:
                time = ReadTime(lines[count])
                if int(time[0]) < MorningPeakHours[0]:
                    count += 1
                else:
                    start = True
            except:
                start = True
                hoursFound = True
                peakHoursM = False
        vehicleMCount = 0
        peakHoursMVehicles = []
        while not hoursFound:
            if count < len(lines):
                time = ReadTime(lines[count])
                if vehicleMCount == 0:
                    initTimeM = time
                if MorningPeakHours[0] <= int(time[0]) <= MorningPeakHours[3]:
                    if vehicleMCount == 0:
                        if MorningPeakHours[0] == int(time[0]):
                            if MorningPeakHours[1] <= int(time[1]):
                                peakHoursMVehicles.append(lines[count])
                                count += 1
                                vehicleMCount += 1
                            else:
                                count += 1
                        else:
                            count += 1
                            vehicleMCount += 1
                    else:
                        peakHoursMVehicles.append(lines[count])
                        count += 1
                        vehicleMCount += 1
                if int(time[0]) >= MorningPeakHours[3] and int(time[1]) >= MorningPeakHours[4] and int(time[2]) >= MorningPeakHours[5]:
                    if MorningPeakHours[3] == int(time[0]) and MorningPeakHours[4] <= int(time[1]):
                        if len(peakHoursMVehicles) != 0:
                            peakHoursMVehicles.pop()
                            vehicleMCount -= 1
                    if len(peakHoursMVehicles) != 0 and len(peakHoursMVehicles) != 1:
                        finalTimeM = ReadTime(peakHoursMVehicles[len(peakHoursMVehicles) - 1])
                    else:
                        peakHoursM = False
                    hoursFound = True
                if int(time[0]) > MorningPeakHours[3]:
                    if len(peakHoursMVehicles) != 0 and len(peakHoursMVehicles) != 1:
                        finalTimeM = ReadTime(peakHoursMVehicles[len(peakHoursMVehicles) - 1])
                    else:
                        peakHoursM = False
                    hoursFound = True
            else:
                if peakHoursM:
                    finalTimeM = ReadTime(peakHoursMVehicles[len(peakHoursMVehicles) - 1])
                hoursFound = True
        MleftVehicles, MrightVehicles, MnullVehicles = CountLeftRight(peakHoursMVehicles)
    if peakHoursN:
        count = 0
        hoursFound = False
        with open("{}/{}".format(folder, file), "r") as f:
            lines = f.readlines()
        for item in lines:
            item = item.replace("\n", "")
            item = item.replace("Vehicle ", "")
            lines[count] = item
            count += 1
        count = 0
        start = False
        while not start:
            try:
                time = ReadTime(lines[count])
                if int(time[0]) < NightPeakHours[0]:
                    count += 1
                else:
                    start = True
            except:
                start = True
                hoursFound = True
                peakHoursN = False
        vehicleNCount = 0
        peakHoursNVehicles = []
        while not hoursFound:
            if count < len(lines):
                time = ReadTime(lines[count])
                if vehicleNCount == 0:
                    initTimeN = time
                if NightPeakHours[0] <= int(time[0]) <= NightPeakHours[3]:
                    if vehicleNCount == 0:
                        if NightPeakHours[0] == int(time[0]):
                            if NightPeakHours[1] <= int(time[1]):
                                peakHoursNVehicles.append(lines[count])
                                count += 1
                                vehicleNCount += 1
                            else:
                                count += 1
                        else:
                            count += 1
                            vehicleNCount += 1
                    else:
                        peakHoursNVehicles.append(lines[count])
                        count += 1
                        vehicleNCount += 1
                if int(time[0]) >= NightPeakHours[3] and int(time[1]) >= NightPeakHours[4] and int(time[2]) >= NightPeakHours[5]:
                    if NightPeakHours[3] == int(time[0]) and NightPeakHours[4] <= int(time[1]):
                        if len(peakHoursNVehicles) != 0:
                            peakHoursNVehicles.pop()
                            vehicleNCount -= 1
                    if len(peakHoursNVehicles) != 0 and len(peakHoursNVehicles) != 1:
                        finalTimeN = ReadTime(peakHoursNVehicles[len(peakHoursNVehicles) - 1])
                    else:
                        peakHoursN = False
                    hoursFound = True
                if int(time[0]) > NightPeakHours[3]:
                    if len(peakHoursNVehicles) != 0 and len(peakHoursNVehicles) != 1:
                        finalTimeN = ReadTime(peakHoursNVehicles[len(peakHoursNVehicles) - 1])
                    else:
                        peakHoursN = False
                    hoursFound = True
            else:
                if peakHoursN:
                    finalTimeN = ReadTime(peakHoursNVehicles[len(peakHoursNVehicles) - 1])
                hoursFound = True
        NleftVehicles, NrightVehicles, NnullVehicles = CountLeftRight(peakHoursNVehicles)
    if not peakHoursM:
        vehicleMCount = "N/A"
        initTimeM = "N/A"
        finalTimeM = "N/A"
        MleftVehicles = "N/A"
        MrightVehicles = "N/A"
        MnullVehicles = "N/A"
    if not peakHoursN:
        vehicleNCount = "N/A"
        initTimeN = "N/A"
        finalTimeN = "N/A"
        NleftVehicles = "N/A"
        NrightVehicles = "N/A"
        NnullVehicles = "N/A"
    return vehicleMCount, initTimeM, finalTimeM, vehicleNCount, initTimeN, finalTimeN, MleftVehicles, MrightVehicles ,MnullVehicles, NleftVehicles, NrightVehicles, NnullVehicles
    


def ReadFile(folder, file):
    
    with open("{}/{}".format(folder, file), "r") as f:
            lines = f.readlines()
    count = 0
    for item in lines:
        item = item.replace("\n", "")
        item = item.replace("Vehicle ", "")
        lines[count] = item
        count += 1
    totalVehicles = count
    totalLeft, totalRight, totalNull = CountLeftRight(lines)
    lineInit = lines[0]
    lineFinal = lines[count - 1]
    timeInit = ReadTime(lineInit)
    timeFinal = ReadTime(lineFinal)

    differenceHour, differenceMinute, differenceSecond, totalHours, vPerHour, vPerHourL, vPerHourR = TimeDifference(timeInit, timeFinal, totalVehicles, totalLeft, totalRight)

    vehicleMCount, initTimeM, finalTimeM, vehicleNCount, initTimeN, finalTimeN ,MleftVehicles, MrightVehicles ,MnullVehicles, NleftVehicles, NrightVehicles, NnullVehicles= PeakHours(timeInit, timeFinal, folder, file)

    MdifferenceHour, MdifferenceMinute, MdifferenceSecond, MtotalHours, MvPerHour, MvPerHourL, MvPerHourR = TimeDifference(initTimeM, finalTimeM, vehicleMCount, MleftVehicles, MrightVehicles)
    NdifferenceHour, NdifferenceMinute, NdifferenceSecond, NtotalHours, NvPerHour, NvPerHourL, NvPerHourR = TimeDifference(initTimeN, finalTimeN, vehicleNCount, NleftVehicles, NrightVehicles)

    wb = load_workbook("DataSpread/{}".format(params.defaultSpreadsheetName))
    ws = wb.active
    currentColumn = 2
    openRow = False
    while not openRow:
        if str(ws.cell(1, currentColumn).value) == "None":
            openRow = True
        else:
            currentColumn += 1
    ws.column_dimensions[map[currentColumn - 1]].width = 15
    ws.cell(1, currentColumn).value = file
    ws.cell(2, currentColumn).value = "{}:{}:{}".format(timeInit[0], timeInit[1], timeInit[2])
    ws.cell(3, currentColumn).value = "{}:{}:{}".format(timeFinal[0], timeFinal[1], timeFinal[2])
    ws.cell(4, currentColumn).value = "{}:{}:{}".format(differenceHour, differenceMinute, differenceSecond)
    ws.cell(5, currentColumn).value = totalVehicles
    ws.cell(6, currentColumn).value = totalLeft
    ws.cell(7, currentColumn).value = totalRight
    ws.cell(8, currentColumn).value = totalNull
    ws.cell(9, currentColumn).value = totalHours
    ws.cell(10, currentColumn).value = vPerHour
    ws.cell(11, currentColumn).value = vPerHourL
    ws.cell(12, currentColumn).value = vPerHourR
    ws.cell(13, currentColumn).value = "{}:{}:{}".format(initTimeM[0], initTimeM[1], initTimeM[2]) if initTimeM != "N/A" else "N/A"
    ws.cell(14, currentColumn).value = "{}:{}:{}".format(finalTimeM[0], finalTimeM[1], finalTimeM[2]) if initTimeM != "N/A" else "N/A"
    ws.cell(15, currentColumn).value = vehicleMCount
    ws.cell(16, currentColumn).value = MtotalHours
    ws.cell(17, currentColumn).value = MvPerHour
    ws.cell(18, currentColumn).value = MleftVehicles
    ws.cell(19, currentColumn).value = MrightVehicles
    ws.cell(20, currentColumn).value = MvPerHourL
    ws.cell(21, currentColumn).value = MvPerHourR
    ws.cell(22, currentColumn).value = "{}:{}:{}".format(initTimeN[0], initTimeN[1], initTimeN[2]) if initTimeN != "N/A" else "N/A"
    ws.cell(23, currentColumn).value = "{}:{}:{}".format(finalTimeN[0], finalTimeN[1], finalTimeN[2]) if initTimeN != "N/A" else "N/A"
    ws.cell(24, currentColumn).value = vehicleNCount
    ws.cell(25, currentColumn).value = NtotalHours
    ws.cell(26, currentColumn).value = NvPerHour
    ws.cell(27, currentColumn).value = NleftVehicles
    ws.cell(28, currentColumn).value = NrightVehicles
    ws.cell(29, currentColumn).value = NvPerHourL
    ws.cell(30, currentColumn).value = NvPerHourR


    wb.save("DataSpread/{}".format(params.defaultSpreadsheetName))
loop = True
while loop:
    actionInput = input("Select Action: ")
    if actionInput == "clear":
        with open("ClearSpreadsheetHztal.py") as f:
            exec(f.read())
    elif actionInput == "new":
        with open("NewSpreadsheetHztal.py") as f:
            exec(f.read())
    elif actionInput == "read":
        folder = params.FileName
        collections = os.listdir("{}/".format(folder))

        msg = "Select File to Read:\n"
        count = 0
        for item in collections:
            msg = msg + "{}: {}".format(count, str(item)) + "\n"
            count += 1
        print(msg)
        readLoop = True
        while readLoop:
            fileInput = input("Select File: ")

            if fileInput == "all":
                count = 0
                for i in collections:
                    ReadFile(folder, collections[count])
                    count += 1
                print("Added a total of {} datasets!".format(count))
                loop = False
            elif fileInput == "end" or fileInput == "":
                readLoop = False
            else:
                file = collections[int(fileInput)]
                ReadFile(folder, file)
                print("Done!")
    elif actionInput == "" or actionInput == "end":
        print("ending!")
        loop = False
    elif actionInput == "help":
        print("Choose one of the following commands:\nread\nclear\nnew\nend")
    else:
        print("Error: no action found!")
        loop = False